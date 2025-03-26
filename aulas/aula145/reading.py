"""Lendo e editando uma planilha de Excel (Workbook e Worksheet)"""

from creating import WORKBOOK_PATH
from pathlib import Path
from openpyxl import Workbook, load_workbook  # type: ignore
from openpyxl.cell import Cell  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore

workbook: Workbook = load_workbook(WORKBOOK_PATH)
worksheet: Worksheet = workbook['My Sheet']

worksheet['B3'].value = 33

row: tuple[Cell]
for row in worksheet.iter_rows(): # type: ignore
    for cell in row:
        print(cell.value, end='\t\t')
    print()
        
workbook.save(WORKBOOK_PATH)
