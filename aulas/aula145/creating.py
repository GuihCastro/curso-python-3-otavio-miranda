"""Criando uma planilha de Excel (Workbook e Worksheet)"""

from pathlib import Path
from openpyxl import Workbook  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore

ROOT_PATH = Path(__file__).parent
WORKBOOK_PATH = ROOT_PATH / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active  # type: ignore
workbook.create_sheet('My Sheet', 0)
worksheet: Worksheet = workbook['My Sheet']
workbook.remove(workbook['Sheet'])

worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    ['Guilherme', 32, 9.5],
    ['Mariana', 32, 10],
    ['Maggie', 9, 9.9],
    ['Kurt', 8, 7.5],
    ['Guts', 5, 5],
    ['Red', 4, 3.75],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
